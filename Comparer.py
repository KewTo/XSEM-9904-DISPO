from glob import glob
import os
from datetime import datetime
from openpyxl import load_workbook
import time
import re

# Find a list of LOTs of XSEM saved into a shared drive and comparing it to a 9904 dispo Excel file that is
# automatically created. The comparison is made to see which is redundant for removal in the 9904 dispo list,
# making the dispo list cleaner


# Set up for the list of folder dates created
def dateRange(createdDate, startDate, endDate):
    createdDate = datetime.strptime(createdDate, '%a %b %d %H:%M:%S %Y')
    startDate = datetime.strptime(startDate, '%d/%m/%Y')
    endDate = datetime.strptime(endDate, '%d/%m/%Y')
    return startDate < createdDate < endDate


# Finding all folders that are between Jan 01 2022 and Sept 11 2022, arbitrary dates, and adding to a list
def get_list_of_folders():
    path = r"Y:\Resist_screening_summaries\X-Section"
    list_of_folders = []
    for filename in os.listdir(path):
        created = time.ctime(os.path.getctime(path + '\\' + filename))
        startDate = '01/01/2022'
        endDate = '11/09/2022'
        if dateRange(created, startDate, endDate):
            list_of_folders.append(filename)
    return list_of_folders


# Find all lots that begin with an IMO and its appropriate six numbers following
def IMO_Lots():
    list_of_IMO = []
    for string in get_list_of_folders():
        if re.match(r'^IMO', string):
            list_of_IMO.append(string[:9])
    return list_of_IMO


# Grab the 9904 dispo Excel list and find all the LOTS adding it into a list
def get_9904_dispo():
    filepath = r'C:\Users\kevinto\Downloads\Search by Module Stores*.xlsx'
    latest_file = max(glob(filepath), key=os.path.getctime)
    wb = load_workbook(latest_file)
    ws = wb.active

    dispo_list = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            dispo_list.append(cell.value)
    return dispo_list


# Compare the list of folders found, and the list of LOTs found in the Excel file; and printing out all in both list
def main():
    print(set(IMO_Lots()) & set(get_9904_dispo()))


if __name__ == '__main__':
    main()
